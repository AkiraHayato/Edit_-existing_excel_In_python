
# =========================================================================
# IMPORTS
# =========================================================================

import xlrd 
from openpyxl import load_workbook

# =========================================================================
# PATHS, OPENS AND LOAD
# =========================================================================

def add_tel(path_compare, path_arv,colunm_ddd,colunm_tel):

    wb_compare = xlrd.open_workbook(path_compare) 
    sheet_compare = wb_compare.sheet_by_index(0)

    wb_arv = xlrd.open_workbook(path_arv) 
    sheet_arv = wb_arv.sheet_by_index(0)

    # =========================================================================

    wb = load_workbook(path_arv)
    ws = wb.active

    # =========================================================================
    # EXTRACT PATHS TO LISTS
    # =========================================================================

    cpf_compare = []
    ddd_compare = []
    tel_compare = []

    cpf_arv = []
    ddd_arv = []
    tel_arv = []

    # =========================================================================

    for i in range(1, sheet_compare.nrows): 
        cpf_compare.append(sheet_compare.cell_value(i, 0))
        ddd_compare.append(sheet_compare.cell_value(i, 2))
        tel_compare.append(sheet_compare.cell_value(i, 3))

    for i in range(1, sheet_arv.nrows): 
        cpf_arv.append(sheet_arv.cell_value(i, 0))
        ddd_arv.append(0)
        tel_arv.append(0)

    # =========================================================================
    # COMPARE METHOD
    # =========================================================================

    count = 0
    count_compare = 0
    count_arv = 0
    
    for c in cpf_compare:   
        for a in cpf_arv:
            if(c == a):
                count+=1
                ddd = ddd_compare[count_compare]
                ddd_arv[count_arv] = ddd
                tel = tel_compare[count_compare]
                tel_arv[count_arv] = tel
                break   
            count_arv+=1  
        if(count == sheet_arv.nrows-1):
            break
        count_arv = 0
        count_compare+= 1
    

    # =========================================================================
    # ADD INFORMATION EXCEL METHOD
    # =========================================================================

    z = 2
    for i in range(sheet_arv.nrows-1):
        _ = ws.cell(column = colunm_ddd,row = z, value = ddd_arv[i])
        _ = ws.cell(column = colunm_tel,row = z, value = tel_arv[i])
        z+=1
    wb.save(path_arv)

if __name__ == "__main__":
    x = 'arv1.xlsx'
    y = 'arv2.xlsx'
    add_tel(x,y,3,4)






